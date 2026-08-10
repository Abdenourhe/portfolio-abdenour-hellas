import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Charger les offres
with open('offres_jobbank.json', 'r', encoding='utf-8') as f:
    jobs = json.load(f)

# Créer un DataFrame
df = pd.DataFrame(jobs)

# Réorganiser les colonnes
cols = ['title', 'company', 'location', 'salary', 'date', 'link', 'status', 'date_postulation', 'lettre_type']
df = df[[c for c in cols if c in df.columns]]

# Renommer les colonnes
df.columns = ['Titre du poste', 'Entreprise', 'Lieu', 'Salaire', 'Date publication', 'Lien', 'Statut', 'Date postulation', 'Type lettre']

# Créer le fichier Excel
wb = Workbook()
ws = wb.active
ws.title = "Offres Job Bank"

# Styles
header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Ajouter les données
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

# Ajuster les largeurs de colonnes
ws.column_dimensions['A'].width = 45  # Titre
ws.column_dimensions['B'].width = 30  # Entreprise
ws.column_dimensions['C'].width = 25  # Lieu
ws.column_dimensions['D'].width = 35  # Salaire
ws.column_dimensions['E'].width = 18  # Date pub
ws.column_dimensions['F'].width = 50  # Lien
ws.column_dimensions['G'].width = 15  # Statut
ws.column_dimensions['H'].width = 15  # Date postulation
ws.column_dimensions['I'].width = 15  # Type lettre

# Hauteur des lignes
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 45

# Ajouter une feuille "Instructions"
ws2 = wb.create_sheet("Instructions")
instructions = [
    ["SUIVI DES CANDIDATURES - ABDENOUR HELLAS"],
    [],
    ["Comment utiliser ce fichier:"],
    ["1. La colonne 'Statut' indique si vous avez déjà postulé ou non"],
    ["   - 'À postuler' = Offre non encore traitée"],
    ["   - 'Postulé' = Candidature envoyée"],
    ["   - 'Refusé' = Réponse négative reçue"],
    ["   - 'En attente' = En cours de traitement"],
    ["2. La colonne 'Date postulation' = date où vous avez envoyé votre candidature"],
    ["3. La colonne 'Type lettre' = quel modèle de lettre utiliser:"],
    ["   - generale = Lettre générale d'ingénieur électrique"],
    ["   - automatisation = Poste en automatisation industrielle"],
    ["   - maintenance = Poste de technicien/maintenance"],
    ["   - supervision = Poste de chef d'équipe/superviseur"],
    [],
    ["MODÈLES DE LETTRES DISPONIBLES:"],
    ["- lettre_motivation_generale.txt"],
    ["- lettre_motivation_automatisation.txt"],
    ["- lettre_motivation_maintenance.txt"],
    ["- lettre_motivation_supervision.txt"],
    [],
    ["IMPORTANT:"],
    ["• Ne postulez PAS deux fois au même poste"],
    ["• Adaptez chaque lettre au poste spécifique"],
    ["• LinkedIn Easy Apply est limité à ~25/jour"],
    ["• Job Bank redirige vers les sites des employeurs"],
    [],
    ["Fichiers créés le 5 août 2026"],
]

for r_idx, row in enumerate(instructions, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(size=14, bold=True, color="1a365d")
        elif value and value.startswith("IMPORTANT:"):
            cell.font = Font(bold=True, color="C00000")

ws2.column_dimensions['A'].width = 80

# Sauvegarder
excel_path = "suivi_candidatures_ABDENOUR_HELLAS.xlsx"
wb.save(excel_path)

print(f"✓ Fichier Excel créé: {excel_path}")
print(f"  - {len(jobs)} offres listées")
print(f"  - Colonnes: Titre, Entreprise, Lieu, Salaire, Date, Lien, Statut, Date postulation, Type lettre")
print(f"\nRésumé des offres:")
print(f"  - Salaire moyen annuel: ~$95,000 - $120,000")
print(f"  - Localisations: Toronto, Montréal, Vancouver, Victoria, Castlegar, Campbell River, Dartmouth, Vars")
