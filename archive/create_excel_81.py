import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Charger les 81 offres
with open('offres_complet_75.json', 'r', encoding='utf-8') as f:
    jobs = json.load(f)

# Créer DataFrame
df = pd.DataFrame(jobs)

# Colonnes
if 'title' in df.columns:
    df = df[['title', 'company', 'location', 'salary', 'date', 'link', 'source', 'status', 'date_postulation', 'lettre_type']]
    df.columns = ['Titre du poste', 'Entreprise', 'Lieu', 'Salaire', 'Date publication', 'Lien', 'Source', 'Statut', 'Date postulation', 'Type lettre']

# Excel
wb = Workbook()
ws = wb.active
ws.title = "Offres - 81 postes"

# Styles
header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# En-têtes + données
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Largeurs
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 55
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 16
ws.column_dimensions['J'].width = 14

# Hauteur
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 40

# Feuille Instructions
ws2 = wb.create_sheet("Instructions")
instructions = [
    ["SUIVI DES CANDIDATURES - ABDENOUR HELLAS"],
    [],
    ["81 OFFRES D'EMPLOI IDENTIFIÉES"],
    ["25 offres Job Bank (originales) + 56 nouvelles offres"],
    [],
    ["CALENDRIER SUGGÉRÉ POUR POSTULER À 50 OFFRES:"],
    ["Semaine 1: Offres 1-15  (3/jour × 5 jours)"],
    ["Semaine 2: Offres 16-30 (3/jour × 5 jours)"],
    ["Semaine 3: Offres 31-45 (3/jour × 5 jours)"],
    ["Semaine 4: Offres 46-50 (relances + nouvelles offres)"],
    [],
    ["POUR POSTULER:"],
    ["1. Ouvrez le fichier lettre_XXX correspondant"],
    ["2. Copiez la lettre personnalisée"],
    ["3. Visitez le lien de l'offre"],
    ["4. Postulez sur le site de l'employeur"],
    ["5. Marquez 'Postulé' dans la colonne Statut"],
    [],
    ["TYPE DE LETTRE:"],
    ["generale = Postes d'ingénieur électrique"],
    ["automatisation = Postes en automatisation"],
    ["maintenance = Postes de technicien/maintenance"],
    ["supervision = Postes de chef d'équipe"],
    [],
    ["IMPORTANT:"],
    ["Ne postulez PAS deux fois au même poste"],
    ["Adaptez 2-3 phrases de chaque lettre au poste spécifique"],
    ["Suivez les entreprises sur LinkedIn avant de postuler"],
    [],
    ["Fichier mis à jour le 5 août 2026"],
]

for r_idx, row in enumerate(instructions, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(size=14, bold=True, color="1a365d")
        elif 'IMPORTANT' in str(value):
            cell.font = Font(bold=True, color="C00000")

ws2.column_dimensions['A'].width = 80

# Sauvegarder
excel_path = "SUIVI_81_OFFRES_ABDENOUR_HELLAS.xlsx"
wb.save(excel_path)

print(f"✓ Fichier Excel créé: {excel_path}")
print(f"  - {len(jobs)} offres listées")
print(f"  - 25 originales Job Bank + 56 nouvelles offres")
print(f"\nRépartition par type de lettre:")
print(df['Type lettre'].value_counts().to_string())
