import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Charger les 254 entreprises
with open('entreprises_200_plus.json', 'r', encoding='utf-8') as f:
    companies = json.load(f)

print(f"Chargement de {len(companies)} entreprises...")

# Créer DataFrame
df = pd.DataFrame(companies)
df = df[['num', 'company', 'sector', 'city', 'poste', 'careers', 'type']]
df.columns = ['#', 'Entreprise', 'Secteur', 'Ville', 'Poste visé', 'Page carrières', 'Type lettre']

# Excel
wb = Workbook()
ws = wb.active
ws.title = "254 Entreprises Canada"

# Styles
header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# En-têtes + données
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Largeurs
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 14

# Hauteur
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 35

# Feuille Instructions
ws2 = wb.create_sheet("Instructions")
instructions = [
    ["MÉGA-LISTE : 254 ENTREPRISES CANADIENNES - ABDENOUR HELLAS"],
    [],
    ["OBJECTIF : Postuler à 200+ postes en génie électrique"],
    [],
    ["MÉTHODE RAPIDE :"],
    ["1. Ouvrez la page carrières de chaque entreprise (colonne F)"],
    ["2. Recherchez 'electrical engineer' ou 'ingénieur électrique' sur leur site"],
    ["3. Postulez directement sur leur portail de carrières"],
    ["4. Marquez 'Postulé' dans le statut"],
    [],
    ["RYTHME RECOMMANDÉ :"],
    ["Semaine 1: Entreprises 1-25  (5/jour × 5 jours)"],
    ["Semaine 2: Entreprises 26-50"],
    ["Semaine 3: Entreprises 51-75"],
    ["Semaine 4: Entreprises 76-100"],
    ["... jusqu'à atteindre 200+ postulations"],
    [],
    ["AVANTAGES DE CETTE MÉTHODE :"],
    ["✓ Moins de concurrence que sur les agrégateurs (Indeed, LinkedIn)"],
    ["✓ Les entreprises recoivent moins de candidatures sur leur site propre"],
    ["✓ Votre CV est vu directement par le service RH"],
    ["✓ Pas de limite journalière comme sur LinkedIn"],
    [],
    ["PRIORITÉS :"],
    ["1. Grandes firmes d'ingénierie (SNC-Lavalin, WSP, Stantec, CIMA+, BBA)"],
    ["2. Utilities (Hydro-Québec, OPG, BC Hydro, Manitoba Hydro)"],
    ["3. Constructeurs électriques (Pomerleau, Aecon, EBC, PCL)"],
    ["4. Mines et alumineries (Rio Tinto, Aluminerie Alouette, Vale)"],
    ["5. Pétrole et gaz (Suncor, Cenovus, Enbridge)"],
    ["6. Manufacturiers (Bombardier, Pratt & Whitney, ABB)"],
    [],
    ["Mise à jour : 5 août 2026"],
]

for r_idx, row in enumerate(instructions, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(size=14, bold=True, color="1a365d")
        elif 'MÉTHODE' in str(value) or 'PRIORITÉS' in str(value) or 'AVANTAGES' in str(value):
            cell.font = Font(bold=True, size=12, color="1a365d")

ws2.column_dimensions['A'].width = 80

# Sauvegarder
excel_path = "MEGA_LISTE_254_ENTREPRISES_ABDENOUR_HELLAS.xlsx"
wb.save(excel_path)

print(f"✓ Fichier Excel créé: {excel_path}")
print(f"  - {len(companies)} entreprises listées")
print(f"\nTop 10 entreprises à cibler en priorité:")
for c in companies[:10]:
    print(f"  {c['num']}. {c['company']} ({c['sector']}) - {c['city']}")
